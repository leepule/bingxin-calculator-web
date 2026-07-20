#!/usr/bin/env python3
"""将 XLSM 的公式、常量、表和命名表达式导出为浏览器计算模型。"""

from __future__ import annotations

import json
import math
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formula import Tokenizer
from openpyxl.worksheet.formula import ArrayFormula


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = ROOT / "spreadsheets" / "冰心计算器Excel版-暗影千机v1.4.xlsm"
DEFAULT_OUTPUT = ROOT / "src" / "generated" / "formula-model.js"


class FormulaParser:
    PRECEDENCE = {
        "=": 1,
        "<>": 1,
        "<": 1,
        ">": 1,
        "<=": 1,
        ">=": 1,
        "&": 2,
        "+": 3,
        "-": 3,
        "*": 4,
        "/": 4,
        "^": 5,
    }

    def __init__(self, formula: str):
        self.tokens = Tokenizer(formula).items
        self.index = 0

    def parse(self) -> list[Any]:
        expression = self._expression(0)
        if self.index != len(self.tokens):
            token = self.tokens[self.index]
            raise ValueError(f"未消费的公式令牌: {token.value} {token.type}/{token.subtype}")
        return expression

    def _expression(self, minimum_precedence: int) -> list[Any]:
        left = self._primary()
        while self.index < len(self.tokens):
            token = self.tokens[self.index]
            if token.type == "OPERATOR-POSTFIX":
                self.index += 1
                left = ["p", token.value, left]
                continue
            if token.type != "OPERATOR-INFIX":
                break
            precedence = self.PRECEDENCE[token.value]
            if precedence < minimum_precedence:
                break
            self.index += 1
            next_minimum = precedence if token.value == "^" else precedence + 1
            right = self._expression(next_minimum)
            left = ["o", token.value, left, right]
        return left

    def _primary(self) -> list[Any]:
        token = self.tokens[self.index]
        self.index += 1

        if token.type == "OPERATOR-PREFIX":
            return ["u", token.value, self._primary()]
        if token.type == "PAREN" and token.subtype == "OPEN":
            expression = self._expression(0)
            self._expect("PAREN", "CLOSE")
            return expression
        if token.type == "FUNC" and token.subtype == "OPEN":
            return self._function(token.value[:-1])
        if token.type == "OPERAND":
            return self._operand(token.value, token.subtype)
        raise ValueError(f"不支持的公式令牌: {token.value} {token.type}/{token.subtype}")

    def _function(self, raw_name: str) -> list[Any]:
        arguments = []
        if self._matches("FUNC", "CLOSE"):
            self.index += 1
            return ["f", normalize_function_name(raw_name), arguments]
        while True:
            if self._matches("SEP", "ARG"):
                arguments.append(["z"])
                self.index += 1
                continue
            arguments.append(self._expression(0))
            if self._matches("SEP", "ARG"):
                self.index += 1
                if self._matches("FUNC", "CLOSE"):
                    arguments.append(["z"])
                continue
            self._expect("FUNC", "CLOSE")
            break
        return ["f", normalize_function_name(raw_name), arguments]

    def _operand(self, raw_value: str, subtype: str) -> list[Any]:
        if subtype == "NUMBER":
            number = float(raw_value)
            return ["n", int(number) if number.is_integer() else number]
        if subtype == "TEXT":
            return ["s", raw_value[1:-1].replace('""', '"')]
        if subtype == "LOGICAL":
            return ["b", raw_value.upper() == "TRUE"]
        return ["r", raw_value]

    def _matches(self, token_type: str, subtype: str) -> bool:
        if self.index >= len(self.tokens):
            return False
        token = self.tokens[self.index]
        return token.type == token_type and token.subtype == subtype

    def _expect(self, token_type: str, subtype: str) -> None:
        if not self._matches(token_type, subtype):
            raise ValueError(f"期望 {token_type}/{subtype}")
        self.index += 1


def normalize_function_name(raw_name: str) -> str:
    name = raw_name.replace("_xlfn.", "").replace("_xlws.", "")
    return name.removeprefix(".").upper() if name.isascii() else name


def formula_text(cell_value: Any) -> tuple[str | None, str | None]:
    if isinstance(cell_value, ArrayFormula):
        return cell_value.text, cell_value.ref
    if isinstance(cell_value, str) and cell_value.startswith("=") and cell_value != "=":
        return cell_value, None
    return None, None


def json_value(cell_value: Any) -> Any:
    if isinstance(cell_value, (datetime, date, time)):
        return cell_value.isoformat()
    if isinstance(cell_value, float) and (math.isnan(cell_value) or math.isinf(cell_value)):
        return None
    return cell_value


def export_sheets(formula_workbook: openpyxl.Workbook, value_workbook: openpyxl.Workbook) -> dict[str, Any]:
    sheets = {}
    for formula_sheet in formula_workbook.worksheets:
        value_sheet = value_workbook[formula_sheet.title]
        coordinates = set(formula_sheet._cells) | set(value_sheet._cells)
        cells = {}
        for row, column in coordinates:
            formula_cell = formula_sheet.cell(row, column)
            cached_value = json_value(value_sheet.cell(row, column).value)
            expression, spill_ref = formula_text(formula_cell.value)
            coordinate = formula_cell.coordinate
            if expression:
                cells[coordinate] = {
                    "f": FormulaParser(expression).parse(),
                    "c": cached_value,
                }
                if spill_ref and spill_ref != coordinate:
                    cells[coordinate]["s"] = spill_ref
            elif formula_cell.value != "=" and formula_cell.value is not None:
                cells[coordinate] = {"v": json_value(formula_cell.value)}
            elif cached_value is not None:
                cells[coordinate] = {"c": cached_value}
        sheets[formula_sheet.title] = {"cells": cells}
    return sheets


def export_tables(workbook: openpyxl.Workbook) -> dict[str, Any]:
    tables = {}
    for worksheet in workbook.worksheets:
        for table in worksheet.tables.values():
            tables[table.name] = {
                "sheet": worksheet.title,
                "ref": table.ref,
                "headers": [column.name for column in table.tableColumns],
                "totals": table.totalsRowCount or 0,
            }
    return tables


def export_names(workbook: openpyxl.Workbook) -> dict[str, Any]:
    names = {}
    for name, definition in workbook.defined_names.items():
        expression = definition.attr_text
        if name.startswith(("_xlpm.", "_xleta.")) or expression == "#NAME?":
            continue
        formula = expression if expression.startswith("=") else f"={expression}"
        names[name] = FormulaParser(formula).parse()
    return names


def export_model(source: Path) -> dict[str, Any]:
    formula_workbook = openpyxl.load_workbook(source, data_only=False, keep_vba=True)
    value_workbook = openpyxl.load_workbook(source, data_only=True, keep_vba=True)
    return {
        "source": source.name,
        "sheets": export_sheets(formula_workbook, value_workbook),
        "tables": export_tables(formula_workbook),
        "names": export_names(formula_workbook),
        "outputs": {
            "dps": ["伤害计算", "AM3"],
            "mainDps": ["计算主页", "G3"],
            "score": ["角色属性", "V13"],
            "spirit": ["角色属性", "V18"],
            "attack": ["角色属性", "V19"],
            "critical": ["角色属性", "V20"],
            "criticalEffect": ["角色属性", "V21"],
            "haste": ["角色属性", "V22"],
            "overcome": ["角色属性", "V23"],
            "strain": ["角色属性", "V24"],
            "surplus": ["角色属性", "V25"],
        },
    }


def formula_model_metrics(model: dict[str, Any]) -> dict[str, int]:
    return {
        "formulaAnchorCount": sum(
            "f" in cell
            for sheet in model["sheets"].values()
            for cell in sheet["cells"].values()
        ),
        "namedExpressionCount": len(model["names"]),
        "sheetCount": len(model["sheets"]),
    }


def main() -> None:
    source = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_SOURCE
    output = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else DEFAULT_OUTPUT
    model = export_model(source)
    output.write_text(
        "window.BINGXIN_FORMULA_MODEL = " + json.dumps(model, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )
    metrics = formula_model_metrics(model)
    print(f"已生成 {output}，公式锚点 {metrics['formulaAnchorCount']} 个，命名表达式 {metrics['namedExpressionCount']} 个。")


if __name__ == "__main__":
    main()
