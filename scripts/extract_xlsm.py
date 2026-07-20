#!/usr/bin/env python3
"""从冰心计算器 XLSM 中提取网页展示所需的缓存数据。"""

from __future__ import annotations

import json
import re
import sys
import zipfile
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl

from export_formula_model import export_model, formula_model_metrics


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = ROOT / "spreadsheets" / "冰心计算器Excel版-暗影千机v1.4.xlsm"
DEFAULT_OUTPUT = ROOT / "src" / "generated" / "data.js"

CUSTOMIZATION_SLOTS = [
    "帽子",
    "衣服",
    "腰带",
    "护手",
    "裤子",
    "鞋子",
    "项链",
    "腰坠",
    "戒指1",
    "戒指2",
    "暗器",
    "武器",
]

UPDATE_DATE_PATTERN = re.compile(r"(?<!\d)(\d{4})[./-](\d{1,2})[./-](\d{1,2})(?!\d)")


def clean(value: Any) -> Any:
    if isinstance(value, float):
        return round(value, 12)
    return value


def percent(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return round(float(value), 12)
    return None


def update_date(update_notes: Any) -> str:
    match = UPDATE_DATE_PATTERN.search(str(update_notes or ""))
    if not match:
        raise ValueError("全局设置!B10 未包含可识别的更新日期")
    year, month, day = (int(part) for part in match.groups())
    return date(year, month, day).isoformat()


def workbook_metrics(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    formula_count = 0
    nonempty_count = 0
    sheets = []

    for worksheet in workbook.worksheets:
        cells = list(worksheet._cells.values())  # 只遍历真实单元格，避开异常的 16384 列范围。
        formulas = sum(
            1
            for cell in cells
            if cell.data_type == "f"
            or (isinstance(cell.value, str) and cell.value.startswith("="))
        )
        nonempty = sum(1 for cell in cells if cell.value is not None)
        formula_count += formulas
        nonempty_count += nonempty
        sheets.append(
            {
                "name": worksheet.title,
                "state": worksheet.sheet_state,
                "rows": worksheet.max_row,
                "columns": worksheet.max_column,
                "formulas": formulas,
                "nonempty": nonempty,
            }
        )

    with zipfile.ZipFile(path) as archive:
        has_vba = any(name.endswith("vbaProject.bin") for name in archive.namelist())

    return {
        "file": path.name,
        "size": path.stat().st_size,
        "sheetCount": len(workbook.sheetnames),
        "visibleSheetCount": sum(sheet["state"] == "visible" for sheet in sheets),
        "definedNameCount": len(workbook.defined_names),
        "formulaCount": formula_count,
        "nonemptyCellCount": nonempty_count,
        "hasVba": has_vba,
        "sheets": sheets,
    }


def extract_equipment_catalog(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    worksheet = workbook["Excel装备"]
    attribute_columns = {
        14: "体质",
        15: "根骨",
        16: "攻击",
        17: "会心",
        18: "会效",
        19: "破防",
        20: "加速",
        21: "破招",
        22: "无双",
        23: "全能",
    }
    items = []

    for row in range(2, worksheet.max_row + 1):
        display_name = worksheet.cell(row, 13).value
        slot = worksheet.cell(row, 3).value
        if not display_name or not slot:
            continue

        attributes = {
            name: clean(worksheet.cell(row, column).value)
            for column, name in attribute_columns.items()
            if worksheet.cell(row, column).value not in (None, "", 0, "0")
        }
        sockets = [
            worksheet.cell(row, column).value
            for column in (24, 25, 26)
            if worksheet.cell(row, column).value not in (None, "", 0, "0")
        ]
        items.append(
            {
                "type": worksheet.cell(row, 1).value,
                "template": worksheet.cell(row, 2).value,
                "slot": slot,
                "name": display_name,
                "setBonus": worksheet.cell(row, 5).value,
                "itemLevel": clean(worksheet.cell(row, 6).value),
                "attributes": attributes,
                "sockets": sockets,
                "refineLevel": clean(worksheet.cell(row, 27).value),
            }
        )

    return items


def extract_build(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    name: str,
    row_start: int,
    stats_label_col: int,
    stats_value_col: int,
    slot_col: int,
    item_col: int,
    detail_cols: tuple[int, int, int, int],
) -> dict[str, Any]:
    stats = {}
    for row in range(row_start, row_start + 9):
        label = worksheet.cell(row, stats_label_col).value
        value = worksheet.cell(row, stats_value_col).value
        if label and value is not None:
            stats[str(label)] = clean(value)

    equipment = []
    for row in range(row_start, row_start + 12):
        slot = worksheet.cell(row, slot_col).value
        item = worksheet.cell(row, item_col).value
        if not slot or not item:
            continue
        details = [
            worksheet.cell(row, column).value
            for column in detail_cols
            if worksheet.cell(row, column).value not in (None, "")
        ]
        equipment.append({"slot": slot, "item": item, "details": details})

    dps = worksheet.cell(row_start + 10, stats_label_col).value
    if not isinstance(dps, (int, float)):
        dps = worksheet.cell(row_start + 10, stats_value_col).value

    return {
        "id": name.lower().replace(" ", "-"),
        "name": name,
        "stats": stats,
        "dps": clean(dps) if isinstance(dps, (int, float)) else None,
        "equipment": equipment,
    }


def extract_build_customization(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    row_start: int,
    detail_cols: tuple[int, int, int, int],
    maximum_refinement: int,
) -> dict[str, Any]:
    major_enchant_col, minor_enchant_col, _, socket_value_col = detail_cols
    customization = {}
    for row_offset, slot in enumerate(CUSTOMIZATION_SLOTS):
        customization[f"{slot}大附魔"] = worksheet.cell(row_start + row_offset, major_enchant_col).value or "无"

    for row_offset, slot in ((0, "帽子"), (1, "衣服"), (2, "腰带"), (3, "护手"), (5, "鞋子")):
        customization[f"{slot}小附魔"] = worksheet.cell(row_start + row_offset, minor_enchant_col).value or "无"

    customization.update(
        {
            "五彩石属性一": worksheet.cell(row_start + 1, socket_value_col).value,
            "五彩石属性二": worksheet.cell(row_start + 2, socket_value_col).value,
            "五彩石属性三": worksheet.cell(row_start + 3, socket_value_col).value,
            "装备孔": clean(worksheet.cell(row_start + 6, socket_value_col).value),
            "精炼等级": maximum_refinement,
        }
    )
    return customization


def extract_builds(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    worksheet = workbook["我的装备"]
    maximum_refinement = int(workbook["计算主页"]["R19"].value)
    layouts = [
        ("装备 1", 3, 2, 3, 4, 5, (8, 9, 10, 11)),
        ("装备 2", 3, 13, 14, 15, 16, (19, 20, 21, 22)),
        ("装备 3", 19, 2, 3, 4, 5, (8, 9, 10, 11)),
        ("装备 4", 19, 13, 14, 15, 16, (19, 20, 21, 22)),
    ]
    builds = []
    for name, row_start, stats_label_col, stats_value_col, slot_col, item_col, detail_cols in layouts:
        build = extract_build(
            worksheet,
            name,
            row_start,
            stats_label_col,
            stats_value_col,
            slot_col,
            item_col,
            detail_cols,
        )
        build["customization"] = extract_build_customization(
            worksheet,
            row_start,
            detail_cols,
            maximum_refinement,
        )
        builds.append(build)
    return builds


def unique_candidates(candidate_values: list[Any]) -> list[Any]:
    return list(dict.fromkeys(candidate for candidate in candidate_values if candidate not in (None, "")))


def counted_column(worksheet: openpyxl.worksheet.worksheet.Worksheet, column: str) -> list[Any]:
    count = int(worksheet[f"{column}2"].value or 0)
    return unique_candidates([worksheet[f"{column}{row}"].value for row in range(3, 3 + count)])


def major_enchantment_fields(references: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    field_specs = [
        ("帽子", "N10", "L"),
        ("衣服", "N11", "Q"),
        ("腰带", "N12", "Q"),
        ("护手", "N13", "M"),
        ("裤子", "N14", "M"),
        ("鞋子", "N15", "L"),
        ("项链", "N16", "R"),
        ("腰坠", "N17", "S"),
        ("戒指1", "N18", "N"),
        ("戒指2", "N19", "N"),
        ("暗器", "N20", "O"),
        ("武器", "N21", "P"),
    ]
    return [
        {
            "key": f"{slot}大附魔",
            "section": "大附魔",
            "label": slot.replace("1", "一").replace("2", "二"),
            "address": address,
            "options": counted_column(references, option_column),
        }
        for slot, address, option_column in field_specs
    ]


def minor_enchantment_fields() -> list[dict[str, Any]]:
    field_specs = [
        ("帽子", "O10", ["无", "伤帽", "小伤帽"]),
        ("衣服", "O11", ["无", "伤衣", "小伤衣"]),
        ("腰带", "O12", ["无", "伤腰", "小伤腰"]),
        ("护手", "O13", ["无", "伤腕", "小伤腕"]),
        ("鞋子", "O15", ["无", "伤鞋", "小伤鞋"]),
    ]
    return [
        {
            "key": f"{slot}小附魔",
            "section": "小附魔",
            "label": slot,
            "address": address,
            "options": options,
        }
        for slot, address, options in field_specs
    ]


def socket_fields(current_data: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    return [
        {
            "key": "五彩石属性一",
            "section": "镶嵌",
            "label": "五彩石 · 属性一",
            "address": "R11",
            "options": unique_candidates([current_data.cell(15, column).value for column in range(19, 28)]),
        },
        {
            "key": "五彩石属性二",
            "section": "镶嵌",
            "label": "五彩石 · 属性二",
            "address": "R12",
            "options": unique_candidates([current_data[f"S{row}"].value for row in range(29, 34)]),
        },
        {
            "key": "五彩石属性三",
            "section": "镶嵌",
            "label": "五彩石 · 属性三",
            "address": "R13",
            "options": unique_candidates([current_data[f"S{row}"].value for row in range(34, 40)]),
        },
        {
            "key": "装备孔",
            "section": "镶嵌",
            "label": "全身装备孔",
            "address": "R16",
            "options": [6, 7, 8],
            "valueType": "number",
        },
        {
            "key": "精炼等级",
            "section": "镶嵌",
            "label": "最高精炼",
            "address": "R19",
            "options": [6, 7, 8],
            "valueType": "number",
        },
    ]


def extract_customization(workbook: openpyxl.Workbook) -> dict[str, Any]:
    home = workbook["计算主页"]
    fields = [
        *major_enchantment_fields(workbook["数据引用"]),
        *minor_enchantment_fields(),
        *socket_fields(workbook["常更新数据"]),
    ]
    return {
        "fields": fields,
        "baseline": {field["key"]: clean(home[field["address"]].value) for field in fields},
    }


def extract_main(workbook: openpyxl.Workbook) -> dict[str, Any]:
    worksheet = workbook["计算主页"]
    score = workbook["角色属性"]["V15"].value
    if not isinstance(score, (int, float)):
        raise ValueError("角色属性!V15 未提供数值装分缓存")
    equipment = []
    for row in range(10, 22):
        slot = worksheet.cell(row, 7).value
        item = worksheet.cell(row, 8).value
        if slot and item:
            details = [
                worksheet.cell(row, column).value
                for column in (14, 15, 17, 18)
                if worksheet.cell(row, column).value not in (None, "")
            ]
            equipment.append({"slot": slot, "item": item, "details": details})

    talent_columns = (7, 10, 11, 12, 13, 14, 16, 17, 18)
    talents = [
        worksheet.cell(28, column).value
        for column in talent_columns
        if worksheet.cell(28, column).value
    ]

    return {
        "title": "冰心诀 · 暗影千机",
        "nickname": worksheet["C33"].value,
        "mainDps": clean(worksheet["G3"].value),
        "score": clean(score),
        "stats": {
            "根骨": clean(worksheet["C11"].value),
            "基础攻击": clean(worksheet["D11"].value),
            "会心": percent(worksheet["C14"].value),
            "会心效果": percent(worksheet["D14"].value),
            "破防": percent(worksheet["C17"].value),
            "无双": percent(worksheet["D17"].value),
            "加速": clean(worksheet["C20"].value),
            "破招": clean(worksheet["D20"].value),
        },
        "settings": {
            "心法": worksheet["G6"].value,
            "循环": worksheet["K3"].value,
            "延迟": worksheet["M3"].value,
            "阵眼": worksheet["K5"].value,
            "数据模式": worksheet["K6"].value,
            "目标等级": worksheet["M5"].value,
            "坦克增益": worksheet["N5"].value,
            "治疗增益": worksheet["N6"].value,
        },
        "consumables": [
            worksheet.cell(row, column).value
            for row in (5, 6)
            for column in (16, 17, 19, 20, 22, 23)
            if worksheet.cell(row, column).value
        ],
        "talents": talents,
        "equipment": equipment,
    }


def extract_damage_analysis(workbook: openpyxl.Workbook) -> dict[str, Any]:
    worksheet = workbook["伤害占比和收益"]
    skills = []
    for row in range(2, 36):
        name = worksheet.cell(row, 5).value
        value = worksheet.cell(row, 6).value
        if name and isinstance(value, (int, float)) and value > 0:
            skills.append({"name": name, "value": clean(value)})

    arrays = []
    for row in range(2, 15):
        name = worksheet.cell(row, 8).value
        value = worksheet.cell(row, 9).value
        if name and isinstance(value, (int, float)):
            arrays.append({"name": name, "value": clean(value)})

    stats = []
    main_sheet = workbook["计算主页"]
    for row in range(2, 9):
        name = worksheet.cell(row, 11).value
        value = worksheet.cell(row, 12).value
        if name and isinstance(value, (int, float)):
            lookup_row = {"会效": 10, "会心": 11, "无双": 12, "破防": 13, "破招": 14, "攻击": 15, "根骨": 16}.get(name)
            unit = main_sheet.cell(lookup_row, 21).value if lookup_row else None
            stats.append({"name": name, "value": clean(value), "unit": clean(unit)})

    return {"skills": skills, "arrays": arrays, "statReturns": stats}


def extract_recommendations(workbook: openpyxl.Workbook) -> dict[str, Any]:
    worksheet = workbook["装备推荐"]
    rows = []
    for row in range(5, 61):
        name = worksheet.cell(row, 2).value
        value = worksheet.cell(row, 5).value
        if name and isinstance(value, (int, float)):
            rows.append({"name": name, "value": clean(value)})
    return {"slot": worksheet["C2"].value, "items": rows}


def extract_haste(workbook: openpyxl.Workbook) -> dict[str, Any]:
    worksheet = workbook["加速表"]
    regular_headers = [worksheet.cell(2, column).value for column in range(2, 10)]
    regular_rows = []
    for row in range(3, 33):
        values = [clean(worksheet.cell(row, column).value) for column in range(2, 10)]
        if any(value is not None for value in values):
            regular_rows.append(dict(zip(regular_headers, values)))

    boundary_rows = []
    for row in range(3, 12):
        boundary = worksheet.cell(row, 13).value
        if boundary is not None:
            boundary_rows.append(
                {
                    "无界玳弦": clean(worksheet.cell(row, 11).value),
                    "无界公共CD": clean(worksheet.cell(row, 12).value),
                    "无界阈值": clean(boundary),
                }
            )
    return {"regular": regular_rows, "boundary": boundary_rows}


def extract_skill_coefficients(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    worksheet = workbook["Excel技能数据"]
    rows = []
    for row in range(2, worksheet.max_row + 1):
        name = worksheet.cell(row, 1).value
        if not name:
            continue
        rows.append(
            {
                "name": name,
                "baseDamage": clean(worksheet.cell(row, 2).value),
                "coefficient": clean(worksheet.cell(row, 3).value),
                "surplusCoefficient": clean(worksheet.cell(row, 4).value),
            }
        )
    return rows


def extract_data(source: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(source, data_only=True, keep_vba=True)
    global_sheet = workbook["全局设置"]
    excel_metrics = workbook_metrics(source)
    formula_metrics = formula_model_metrics(export_model(source))
    update_notes = global_sheet["B10"].value

    wps_path = ROOT / "spreadsheets" / "冰心计算器WPS版-暗影千机v1.4.xlsm"
    source_comparison = [excel_metrics]
    if wps_path.exists() and wps_path != source:
        source_comparison.append(workbook_metrics(wps_path))

    return {
        "meta": {
            "version": global_sheet["A2"].value,
            "updatedAt": update_date(update_notes),
            "updateNotes": update_notes,
            "source": source.name,
            "formulaModel": formula_metrics,
            "sourceComparison": source_comparison,
            "reference": {
                "title": "JX3BOX · 原始 Excel 参考页面",
                "url": "https://www.jx3box.com/bps/107327",
            },
            "notice": "浏览器不直接执行 vbaProject.bin；读取、保存配装和装备推荐已用 JavaScript 等价实现，自定义配装由网页公式引擎执行完整公式链重算。",
        },
        "main": extract_main(workbook),
        "builds": extract_builds(workbook),
        "customization": extract_customization(workbook),
        "analysis": extract_damage_analysis(workbook),
        "recommendations": extract_recommendations(workbook),
        "haste": extract_haste(workbook),
        "equipmentCatalog": extract_equipment_catalog(workbook),
        "skillCoefficients": extract_skill_coefficients(workbook),
    }


def main() -> None:
    source = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_SOURCE
    output = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else DEFAULT_OUTPUT
    data = extract_data(source)
    output.write_text(
        "window.APP_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"已生成 {output}，装备 {len(data['equipmentCatalog'])} 件。")


if __name__ == "__main__":
    main()
